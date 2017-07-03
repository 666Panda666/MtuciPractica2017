import openpyxl as op

f_out = open('file.txt', 'r')
wb1 = op.load_workbook(f_out.readline()[:-1])
mat_ex = wb1.active
wb2 = op.load_workbook(f_out.readline()[:-1])
time_of_work = wb1.get_active_sheet()
wb3 = op.load_workbook(f_out.readline()[:-1])
amount = wb1.get_active_sheet()

wb4 = op.load_workbook(f_out.readline())
req = wb4.active
mat = [[None]*(mat_ex.max_column-1) for i in range(mat_ex.max_row-1)]
for j in range(2,mat_ex.max_column + 1):
    for i in range(2,mat_ex.max_row + 1):
        mat[i-2][j-2] = mat_ex.cell(row = i, column = j).value
req_list = []
for i in range(1, req.max_column+1):
    req_list.append(req.cell(row = 1, column = i).value)
obj_list = []
for i in range(2, mat_ex.max_column+1):
    obj_list.append(mat_ex.cell(row = 1, column = i).value)
ind = []
for i in range(0,len(obj_list)):
   for j in range(0,len(req_list)):
       if obj_list[i].startswith(req_list[j]):
           ind.append(i)
new_mat = [[None]*len(ind) for i in range(len(ind))]
for i in range(0,len(ind)):
    for j in range(0,len(ind)):
        new_mat[i][j] = mat[ind[i]][ind[j]]


